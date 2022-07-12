from django.conf import settings
from django.urls import reverse_lazy
import openpyxl
import pandas as pd
import os


def file_is_exist(path):
    result = os.path.exists(path)
    
    return result


def get_current_file(path):
    file = openpyxl.load_workbook(path)
    return file


def get_current_sheet(path):
    file = openpyxl.load_workbook(path)
    current_sheet = file.active

    return current_sheet


def get_excel_data(path):
    try:
        data = pd.read_excel(path)
        data = data.fillna(value="Empty")
        records = data.to_dict("index")
        keys = [key for key in records][1:]

        for key in keys:
            records[key]["index"] = key
            records[key]["details"] = reverse_lazy("ExcelDetailAPIView", args=[key])

        return records

    except Exception as e:
        data = None
        return data


def search_in_column(worksheet, search_string, column="A"):
    
    for row in range(1, worksheet.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if worksheet[coordinate].value == search_string:
            
            return row
    return None